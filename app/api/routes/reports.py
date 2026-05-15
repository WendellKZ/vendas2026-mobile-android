from datetime import datetime, date
from io import BytesIO
from typing import Any

from fastapi import APIRouter, Depends, Query, Request
from fastapi.responses import RedirectResponse, StreamingResponse
from sqlalchemy import func, case
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from app.auth import current_user_from_cookie, is_admin
from app.db.deps import get_db
from app.models.customer import Customer
from app.models.order import Order, OrderItem
from app.models.product import Product
from app.models.user import User

web_router = APIRouter()

REPORTS = {
    "vendas_periodo": "Vendas por período",
    "ranking_vendedores": "Ranking de vendedores",
    "produtos_vendidos": "Produtos mais vendidos",
    "clientes_compram": "Clientes que mais compram",
    "pedidos_status": "Pedidos por status",
    "descontos": "Descontos concedidos",
    "comissoes": "Comissões",
    "personalizado": "Relatório personalizado",
}

CUSTOM_GROUPS = {
    "vendedor": "Vendedor",
    "cliente": "Cliente",
    "produto": "Produto",
    "status": "Status do pedido",
    "pagamento": "Condição de pagamento",
    "mes": "Mês",
}


def _parse_date(value: str | None):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d")
    except ValueError:
        return None


def _money(value):
    return float(value or 0)


def _fmt_money(value):
    return f"R$ {_money(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _ctx(request: Request, db: Session):
    user = current_user_from_cookie(request, db)
    return {
        "active_seller": user,
        "current_user": user,
        "is_admin": is_admin(user),
        "sellers_menu": db.query(User).filter(User.active == True).order_by(User.name).all(),  # noqa: E712
    }


def _require_user(request: Request, db: Session):
    user = current_user_from_cookie(request, db)
    if not user:
        return None
    return user


def _base_order_query(db: Session, user: User, data_inicio: str | None, data_fim: str | None, seller_id: str | None, customer_id: str | None, status: str | None):
    query = db.query(Order)
    if user.role != "admin":
        query = query.filter(Order.seller_id == user.id)
    elif seller_id:
        query = query.filter(Order.seller_id == int(seller_id))
    if customer_id:
        query = query.filter(Order.customer_id == int(customer_id))
    if status:
        query = query.filter(Order.status == status)
    start = _parse_date(data_inicio)
    end = _parse_date(data_fim)
    if start:
        query = query.filter(Order.created_at >= start)
    if end:
        query = query.filter(Order.created_at < datetime(end.year, end.month, end.day, 23, 59, 59))
    return query


def _filter_order_ids(base_query):
    return [row[0] for row in base_query.with_entities(Order.id).all()]


def _summary(base_query):
    row = base_query.with_entities(
        func.count(Order.id),
        func.coalesce(func.sum(Order.total_net), 0),
        func.coalesce(func.sum(Order.total_gross), 0),
        func.coalesce(func.sum(Order.total_discount), 0),
        func.coalesce(func.sum(Order.commission_total), 0),
    ).one()
    pedidos = int(row[0] or 0)
    total = _money(row[1])
    return {
        "pedidos": pedidos,
        "faturamento": total,
        "bruto": _money(row[2]),
        "desconto": _money(row[3]),
        "comissao": _money(row[4]),
        "ticket_medio": total / pedidos if pedidos else 0,
    }


def _rows_for_report(db: Session, base_query, report_type: str, custom_group: str):
    order_ids = _filter_order_ids(base_query)
    if not order_ids:
        return [], []

    if report_type == "ranking_vendedores":
        rows = db.query(User.name, func.count(Order.id), func.coalesce(func.sum(Order.total_net), 0), func.coalesce(func.sum(Order.commission_total), 0)).join(Order, Order.seller_id == User.id).filter(Order.id.in_(order_ids)).group_by(User.id, User.name).order_by(func.sum(Order.total_net).desc()).all()
        return ["Vendedor", "Pedidos", "Total vendido", "Comissão"], [[r[0] or "Sem vendedor", r[1], _fmt_money(r[2]), _fmt_money(r[3])] for r in rows]

    if report_type == "produtos_vendidos":
        rows = db.query(Product.name, Product.category, func.coalesce(func.sum(OrderItem.quantity), 0), func.coalesce(func.sum(OrderItem.total), 0)).join(OrderItem, OrderItem.product_id == Product.id).filter(OrderItem.order_id.in_(order_ids)).group_by(Product.id, Product.name, Product.category).order_by(func.sum(OrderItem.total).desc()).all()
        return ["Produto", "Categoria", "Quantidade", "Total vendido"], [[r[0], r[1] or "-", f"{float(r[2] or 0):,.0f}".replace(",", "."), _fmt_money(r[3])] for r in rows]

    if report_type == "clientes_compram":
        rows = db.query(Customer.name, Customer.document, Customer.city, func.count(Order.id), func.coalesce(func.sum(Order.total_net), 0), func.coalesce(func.avg(Order.total_net), 0)).join(Order, Order.customer_id == Customer.id).filter(Order.id.in_(order_ids)).group_by(Customer.id, Customer.name, Customer.document, Customer.city).order_by(func.sum(Order.total_net).desc()).all()
        return ["Cliente", "CNPJ/CPF", "Cidade", "Pedidos", "Total comprado", "Ticket médio"], [[r[0], r[1] or "-", r[2] or "-", r[3], _fmt_money(r[4]), _fmt_money(r[5])] for r in rows]

    if report_type == "pedidos_status":
        rows = db.query(Order.status, func.count(Order.id), func.coalesce(func.sum(Order.total_net), 0)).filter(Order.id.in_(order_ids)).group_by(Order.status).order_by(func.count(Order.id).desc()).all()
        return ["Status", "Pedidos", "Total"], [[r[0] or "-", r[1], _fmt_money(r[2])] for r in rows]

    if report_type == "descontos":
        rows = db.query(User.name, func.count(Order.id), func.coalesce(func.sum(Order.total_gross), 0), func.coalesce(func.sum(Order.total_discount), 0), func.coalesce(func.avg(Order.max_discount_applied), 0)).join(User, User.id == Order.seller_id, isouter=True).filter(Order.id.in_(order_ids)).group_by(User.id, User.name).order_by(func.sum(Order.total_discount).desc()).all()
        return ["Vendedor", "Pedidos", "Total bruto", "Desconto concedido", "Desconto médio %"], [[r[0] or "Sem vendedor", r[1], _fmt_money(r[2]), f"{float(r[3] or 0):.2f}%".replace(".", ","), _fmt_money(r[4])] for r in rows]

    if report_type == "comissoes":
        rows = db.query(User.name, func.count(Order.id), func.coalesce(func.sum(Order.total_net), 0), func.coalesce(func.avg(Order.commission_percent), 0), func.coalesce(func.sum(Order.commission_total), 0)).join(User, User.id == Order.seller_id, isouter=True).filter(Order.id.in_(order_ids)).group_by(User.id, User.name).order_by(func.sum(Order.commission_total).desc()).all()
        return ["Vendedor", "Pedidos", "Total vendido", "Comissão média %", "Comissão total"], [[r[0] or "Sem vendedor", r[1], _fmt_money(r[2]), f"{float(r[3] or 0):.2f}%".replace(".", ","), _fmt_money(r[4])] for r in rows]

    if report_type == "personalizado":
        if custom_group == "cliente":
            rows = db.query(Customer.name, func.count(Order.id), func.coalesce(func.sum(Order.total_net), 0), func.coalesce(func.sum(Order.total_discount), 0), func.coalesce(func.sum(Order.commission_total), 0)).join(Order, Order.customer_id == Customer.id).filter(Order.id.in_(order_ids)).group_by(Customer.id, Customer.name).order_by(func.sum(Order.total_net).desc()).all()
            return ["Cliente", "Pedidos", "Total vendido", "Desconto", "Comissão"], [[r[0], r[1], _fmt_money(r[2]), _fmt_money(r[3]), _fmt_money(r[4])] for r in rows]
        if custom_group == "produto":
            rows = db.query(Product.name, func.coalesce(func.sum(OrderItem.quantity), 0), func.coalesce(func.sum(OrderItem.total), 0)).join(OrderItem, OrderItem.product_id == Product.id).filter(OrderItem.order_id.in_(order_ids)).group_by(Product.id, Product.name).order_by(func.sum(OrderItem.total).desc()).all()
            return ["Produto", "Quantidade", "Total vendido"], [[r[0], f"{float(r[1] or 0):,.0f}".replace(",", "."), _fmt_money(r[2])] for r in rows]
        if custom_group == "status":
            rows = db.query(Order.status, func.count(Order.id), func.coalesce(func.sum(Order.total_net), 0)).filter(Order.id.in_(order_ids)).group_by(Order.status).order_by(func.count(Order.id).desc()).all()
            return ["Status", "Pedidos", "Total vendido"], [[r[0] or "-", r[1], _fmt_money(r[2])] for r in rows]
        if custom_group == "pagamento":
            rows = db.query(Order.payment_condition, func.count(Order.id), func.coalesce(func.sum(Order.total_net), 0)).filter(Order.id.in_(order_ids)).group_by(Order.payment_condition).order_by(func.sum(Order.total_net).desc()).all()
            return ["Condição de pagamento", "Pedidos", "Total vendido"], [[r[0] or "-", r[1], _fmt_money(r[2])] for r in rows]
        if custom_group == "mes":
            dialect = db.bind.dialect.name
            month_expr = func.to_char(Order.created_at, "YYYY-MM") if dialect == "postgresql" else func.strftime("%Y-%m", Order.created_at)
            rows = db.query(month_expr.label("mes"), func.count(Order.id), func.coalesce(func.sum(Order.total_net), 0), func.coalesce(func.sum(Order.commission_total), 0)).filter(Order.id.in_(order_ids)).group_by(month_expr).order_by(month_expr).all()
            return ["Mês", "Pedidos", "Total vendido", "Comissão"], [[r[0] or "Sem data", r[1], _fmt_money(r[2]), _fmt_money(r[3])] for r in rows]
        report_type = "ranking_vendedores"

    rows = db.query(Order.id, Order.created_at, Customer.name, User.name, Order.status, Order.payment_condition, Order.total_gross, Order.total_discount, Order.total_net, Order.commission_total).join(Customer, Customer.id == Order.customer_id, isouter=True).join(User, User.id == Order.seller_id, isouter=True).filter(Order.id.in_(order_ids)).order_by(Order.id.desc()).all()
    return ["Pedido", "Data", "Cliente", "Vendedor", "Status", "Pagamento", "Bruto", "Desconto", "Total", "Comissão"], [[r[0], r[1].strftime("%d/%m/%Y") if r[1] else "-", r[2] or "-", r[3] or "-", r[4] or "-", r[5] or "-", _fmt_money(r[6]), _fmt_money(r[7]), _fmt_money(r[8]), _fmt_money(r[9])] for r in rows]


def _build_workbook(title: str, filters: dict[str, Any], summary: dict[str, Any], headers: list[str], rows: list[list[Any]]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio"
    accent = "1E3A8A"
    light = "EEF5FF"
    line = "D9E4F2"
    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True, color=accent)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 6))
    ws["A2"] = "Gerado em"
    ws["B2"] = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws["A3"] = "Filtros"
    ws["B3"] = " | ".join([f"{k}: {v}" for k, v in filters.items() if v]) or "Sem filtros específicos"
    ws["A5"] = "Pedidos"
    ws["B5"] = summary["pedidos"]
    ws["C5"] = "Faturamento"
    ws["D5"] = summary["faturamento"]
    ws["E5"] = "Ticket médio"
    ws["F5"] = summary["ticket_medio"]
    ws["G5"] = "Desconto"
    ws["H5"] = summary["desconto"]
    ws["I5"] = "Comissão"
    ws["J5"] = summary["comissao"]
    for cell in ws[5]:
        cell.fill = PatternFill("solid", fgColor=light)
        cell.font = Font(bold=True, color="0F172A")
        cell.alignment = Alignment(horizontal="center")
    start = 7
    for col, header in enumerate(headers, 1):
        cell = ws.cell(start, col, header)
        cell.fill = PatternFill("solid", fgColor=accent)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    for r_idx, row in enumerate(rows, start + 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(r_idx, c_idx, value)
            if isinstance(value, float):
                cell.number_format = 'R$ #,##0.00'
    thin = Side(style="thin", color=line)
    max_row = start + max(len(rows), 1)
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max(len(headers), 10)):
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="center")
    widths = [16, 16, 28, 24, 18, 22, 16, 16, 16, 16]
    for idx in range(1, max(len(headers), 10) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = widths[idx - 1] if idx <= len(widths) else 18
    ws.freeze_panes = "A8"
    return wb


@web_router.get("/relatorios")
def relatorios_page(
    request: Request,
    db: Session = Depends(get_db),
    tipo: str = Query("vendas_periodo"),
    data_inicio: str | None = Query(None),
    data_fim: str | None = Query(None),
    seller_id: str | None = Query(None),
    customer_id: str | None = Query(None),
    status: str | None = Query(None),
    agrupamento: str = Query("vendedor"),
):
    user = _require_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    if tipo not in REPORTS:
        tipo = "vendas_periodo"
    if agrupamento not in CUSTOM_GROUPS:
        agrupamento = "vendedor"
    base_query = _base_order_query(db, user, data_inicio, data_fim, seller_id, customer_id, status)
    summary = _summary(base_query)
    headers, rows = _rows_for_report(db, base_query, tipo, agrupamento)
    sellers = db.query(User).filter(User.active == True).order_by(User.name).all() if user.role == "admin" else [user]  # noqa: E712
    customers = db.query(Customer).order_by(Customer.name).limit(500).all()
    statuses = [r[0] for r in db.query(Order.status).filter(Order.status.isnot(None)).distinct().order_by(Order.status).all()]
    return request.app.state.templates.TemplateResponse("relatorios.html", {
        "request": request,
        "title": "Relatórios Comerciais",
        "subtitle": "Filtros flexíveis, indicadores essenciais e exportação em Excel.",
        "reports": REPORTS,
        "custom_groups": CUSTOM_GROUPS,
        "selected_type": tipo,
        "selected_group": agrupamento,
        "summary": summary,
        "headers": headers,
        "rows": rows,
        "sellers": sellers,
        "customers": customers,
        "statuses": statuses,
        "filters": {"data_inicio": data_inicio or "", "data_fim": data_fim or "", "seller_id": seller_id or "", "customer_id": customer_id or "", "status": status or ""},
        **_ctx(request, db),
    })


@web_router.get("/relatorios/exportar")
def relatorios_exportar(
    request: Request,
    db: Session = Depends(get_db),
    tipo: str = Query("vendas_periodo"),
    data_inicio: str | None = Query(None),
    data_fim: str | None = Query(None),
    seller_id: str | None = Query(None),
    customer_id: str | None = Query(None),
    status: str | None = Query(None),
    agrupamento: str = Query("vendedor"),
):
    user = _require_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    if tipo not in REPORTS:
        tipo = "vendas_periodo"
    if agrupamento not in CUSTOM_GROUPS:
        agrupamento = "vendedor"
    base_query = _base_order_query(db, user, data_inicio, data_fim, seller_id, customer_id, status)
    summary = _summary(base_query)
    headers, rows = _rows_for_report(db, base_query, tipo, agrupamento)
    filters = {"Período inicial": data_inicio, "Período final": data_fim, "Vendedor": seller_id, "Cliente": customer_id, "Status": status, "Agrupamento": CUSTOM_GROUPS.get(agrupamento)}
    wb = _build_workbook(REPORTS[tipo], filters, summary, headers, rows)
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"relatorio_{tipo}_{date.today().isoformat()}.xlsx"
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})
